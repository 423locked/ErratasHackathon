from openpyxl import load_workbook # EXCEL

# Обработка расписания из файла
def get_lessons(NAME):

	lessons = [["" for i in range(7)] for i in range(14)] # 7 пар каждый день в течение двух недель

	excel_file = load_workbook('./IKB.xlsx')

	sheet_names = excel_file.sheetnames

	for sheet_name in sheet_names:

		sheet = excel_file[sheet_name]

		try:
			_ = sheet["DG2"].value
		except:
			continue

		if sheet["DG2"].value == NAME or sheet["DV2"].value == NAME or sheet["DL2"].value == NAME: # Поиск группы

			if sheet["DL2"].value == NAME:
				row = "DL"
			elif sheet["DG2"].value == NAME:
				row = "DG"
			else:
				row = "DV"


			for i in range(4,88): # Строки таблицы с 4-ой по 88-ую
				# В четных ячейках лежат пары нечетных недель и наоборот
				day = (i-4)//12 + (i%2)*7 # День пары
				pare_index = ((i-4)%14)//2 # Номер пары 				

				if sheet[row+str(i)].value != "" and not (sheet[row+str(i)].value is None): # Если пара существует
					lessons[day][pare_index] = [sheet[row+str(i)].value,sheet[row[:-1]+chr(ord(row[-1])+1)+str(i)].value,sheet[row[:-1]+chr(ord(row[-1])+2)+str(i)].value,sheet[row[:-1]+chr(ord(row[-1])+3)+str(i)].value]
 	
					# Делаем вывод красивым и понятным

					if len(lessons[day][pare_index][0].split("\n"))>1: # удаление повторов и разъединение двойных пар

						for i in range(0,4):
							if lessons[day][pare_index][i].split("\n")[0]!=lessons[day][pare_index][i].split("\n")[1]:
								lessons[day][pare_index][i] = " или ".join(lessons[day][pare_index][i].split("\n"))
							else:
								lessons[day][pare_index][i] = lessons[day][pare_index][i].split('\n')[0]

					# Удаление доп иформации о паре	
					if "кр." in lessons[day][pare_index][0]:
						lessons[day][pare_index][0] = " ".join(lessons[day][pare_index][0].split(" ")[3:])

					if ord(lessons[day][pare_index][0][0]) < 1040:

						lessons[day][pare_index][0] = " ".join(lessons[day][pare_index][0].split(" ")[2:])

					# Обработка здания
					place = lessons[day][pare_index][3].split(" ")
					for i in range(len(place)):
						if place[i] == "(С-20)":
							place[i] = "на Стромынке"
						elif place[i] == "(В-78)":
							place[i] = "на Вернадке 78"
						elif place[i] == "(МП-1)":
							place[i] = "на Пироговке"
					lessons[day][pare_index][3] = " ".join(place)
			break

	# Добавляем воскресения, чтобы можно было работать с датами в семидневных неделях			
	lessons.insert(6,[])
	lessons.append([])
	#print(lessons)
	return lessons
'''
Возвращает массив из 14 массивов (14 дней в двух неделях)
В каждом из этих массивов 7 массивов с информацией о паре(7 пар в день)
